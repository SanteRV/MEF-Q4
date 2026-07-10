"""Exportador a Excel: una hoja por paso, con tablas y matrices."""
from __future__ import annotations
from pathlib import Path
import re

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..fem.steps import Procedure


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=14, bold=True, color="1F4E78")
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _safe_sheet_name(name: str) -> str:
    """Excel: máximo 31 caracteres, no permite : \\ / ? * [ ]"""
    name = re.sub(r"[:\\/?*\[\]]", " ", name)
    return name[:31]


# Excel solo guarda hasta 15 dígitos significativos (limite de float64).
# El value almacenado en la celda SIEMPRE es float64 completo; estos formatos
# controlan solo cómo se MUESTRA en pantalla.
NUM_FMT = "0.00000000000000;-0.00000000000000;0"    # 14 decimales
SCI_FMT = "0.00000000000000E+00"                    # 14 decimales en científico


def _write_dataframe(ws, df, start_row: int) -> int:
    """Escribe un DataFrame y devuelve la fila siguiente al final."""
    for c, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=c, value=str(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for r, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for c, val in enumerate(row.values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            if isinstance(val, float):
                # Notación científica solo para valores extremos (legibilidad)
                if val != 0 and (abs(val) < 1e-4 or abs(val) > 1e6):
                    cell.number_format = SCI_FMT
                else:
                    cell.number_format = NUM_FMT
    return start_row + len(df) + 2


def _write_matrix(ws, name: str, M: np.ndarray, start_row: int) -> int:
    ws.cell(row=start_row, column=1, value=name).font = Font(bold=True)
    start_row += 1
    if M.ndim == 1:
        M = M.reshape(-1, 1)
    for r in range(M.shape[0]):
        for c in range(M.shape[1]):
            v = float(M[r, c])
            cell = ws.cell(row=start_row + r, column=1 + c, value=v)
            cell.border = BORDER
            if v != 0 and (abs(v) < 1e-4 or abs(v) > 1e6):
                cell.number_format = SCI_FMT
            else:
                cell.number_format = NUM_FMT
    return start_row + M.shape[0] + 2


def export_to_excel_organized(procedure: Procedure, path: Path) -> None:
    """Exporta a Excel con hojas categorizadas para procesamiento posterior.

    Estructura del archivo (corrección E16 del documento):
        Hoja 1: Resumen — descripción del análisis + lista de pasos
        Hoja 2: Datos_entrada — Nodos, Elementos, Propiedades (1 nodo por fila)
        Hoja 3: Apoyos_y_cargas — Solo BC y cargas (1 línea por nodo)
        Hoja 4: Resultados_nodales — Desplazamientos y reacciones por nodo
        Hoja 5: Esfuerzos_GP — Tabla larga: elemento, GP, εx, εy, γxy, σx, σy, τxy
        Hoja 6: Esfuerzos_esquinas — Idem en las 4 esquinas/nodos
        Hojas siguientes: una por matriz importante (K_global, K^e, D, B, F)

    Cada hoja tiene la información en formato tabular limpio listo para
    importar en MATLAB, Python, R o post-procesar en Excel mismo.
    """
    import pandas as pd

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # ===== Hoja 1: Resumen =====
    ws = wb.create_sheet("Resumen")
    ws.cell(row=1, column=1, value="Análisis MEF Q4 — Resumen").font = TITLE_FONT
    ws.cell(row=2, column=1,
            value="Archivo organizado por categorías para post-proceso.").alignment = Alignment(
        wrap_text=True
    )
    ws.cell(row=4, column=1, value="Paso").font = HEADER_FONT
    ws.cell(row=4, column=2, value="Título").font = HEADER_FONT
    ws.cell(row=4, column=3, value="Contenido").font = HEADER_FONT
    for i, st in enumerate(procedure.steps, start=5):
        ws.cell(row=i, column=1, value=i - 4)
        ws.cell(row=i, column=2, value=st.title)
        items = []
        if st.tables:
            items.append(f"{len(st.tables)} tabla(s)")
        if st.matrices:
            items.append(f"{len(st.matrices)} matriz(ces)")
        if st.plot_key:
            items.append(f"gráfica '{st.plot_key}'")
        ws.cell(row=i, column=3, value=", ".join(items) or "—")
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40

    result = procedure.result
    # Encontrar pasos clave por su título (más robusto que por índice)
    def _step_by_keyword(keyword: str):
        for st in procedure.steps:
            if keyword.lower() in st.title.lower():
                return st
        return None

    # ===== Hoja 2: Datos_entrada =====
    step1 = _step_by_keyword("Datos de entrada")
    if step1 is not None and step1.tables:
        ws = wb.create_sheet("Datos_entrada")
        ws.cell(row=1, column=1, value="Datos de entrada del modelo").font = TITLE_FONT
        row = 3
        for name, df in step1.tables.items():
            ws.cell(row=row, column=1, value=name).font = Font(bold=True, color="1F4E78")
            row = _write_dataframe(ws, df, row + 1)
            row += 1

    # ===== Hoja 3: Apoyos_y_cargas =====
    step2 = _step_by_keyword("Condiciones de borde")
    if step2 is not None and step2.tables:
        ws = wb.create_sheet("Apoyos_y_cargas")
        ws.cell(row=1, column=1, value="Condiciones de borde y cargas").font = TITLE_FONT
        row = 3
        for name, df in step2.tables.items():
            ws.cell(row=row, column=1, value=name).font = Font(bold=True, color="1F4E78")
            row = _write_dataframe(ws, df, row + 1)
            row += 1

    # ===== Hoja 4: Resultados_nodales =====
    if result is not None:
        ws = wb.create_sheet("Resultados_nodales")
        ws.cell(row=1, column=1, value="Resultados nodales").font = TITLE_FONT
        # Construir DataFrame combinado: nodo, ux, uy, Rx, Ry
        # Necesito structure; lo extraigo del primer paso plot_key="mesh"
        # Hack: usar el procedure mismo; los desplazamientos están en result
        ws.cell(row=3, column=1, value="Nodo").font = HEADER_FONT
        ws.cell(row=3, column=2, value="ux (m)").font = HEADER_FONT
        ws.cell(row=3, column=3, value="uy (m)").font = HEADER_FONT
        ws.cell(row=3, column=4, value="Rx (N)").font = HEADER_FONT
        ws.cell(row=3, column=5, value="Ry (N)").font = HEADER_FONT
        # Inferir nodos del paso 1
        n_nodes = len(result.displacements) // 2
        for i in range(n_nodes):
            r = 4 + i
            ws.cell(row=r, column=1, value=i + 1)
            ws.cell(row=r, column=2, value=float(result.displacements[2 * i]))
            ws.cell(row=r, column=3, value=float(result.displacements[2 * i + 1]))
            ws.cell(row=r, column=4, value=float(result.reactions[2 * i]))
            ws.cell(row=r, column=5, value=float(result.reactions[2 * i + 1]))
            for c in range(1, 6):
                cell = ws.cell(row=r, column=c)
                cell.border = BORDER
                if c > 1:
                    cell.number_format = NUM_FMT
        for c in range(1, 6):
            ws.cell(row=3, column=c).fill = HEADER_FILL
            ws.cell(row=3, column=c).border = BORDER
        for col_letter in "ABCDE":
            ws.column_dimensions[col_letter].width = 16

    # ===== Hoja 5: Esfuerzos_GP =====
    if result is not None and result.elements:
        ws = wb.create_sheet("Esfuerzos_GP")
        ws.cell(row=1, column=1,
                value="Deformaciones y esfuerzos en los puntos de Gauss").font = TITLE_FONT
        headers = ["Elemento", "Punto", "εx", "εy", "γxy",
                   "σx (Pa)", "σy (Pa)", "τxy (Pa)"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
        r = 4
        for el_idx, el_res in enumerate(result.elements, start=1):
            for gp_idx, (eps, sig) in enumerate(
                zip(el_res.strains_at_gauss, el_res.stresses_at_gauss), start=1
            ):
                ws.cell(row=r, column=1, value=el_idx).border = BORDER
                ws.cell(row=r, column=2, value=f"GP{gp_idx}").border = BORDER
                for c, v in enumerate(
                    [eps[0], eps[1], eps[2], sig[0], sig[1], sig[2]], start=3
                ):
                    cell = ws.cell(row=r, column=c, value=float(v))
                    cell.number_format = NUM_FMT
                    cell.border = BORDER
                r += 1
        for col_letter in "ABCDEFGH":
            ws.column_dimensions[col_letter].width = 16

    # ===== Hoja 6: Esfuerzos_esquinas =====
    if result is not None and result.elements:
        ws = wb.create_sheet("Esfuerzos_esquinas")
        ws.cell(row=1, column=1,
                value="Esfuerzos en las esquinas (nodos)").font = TITLE_FONT
        headers = ["Elemento", "Esquina", "εx", "εy", "γxy",
                   "σx (Pa)", "σy (Pa)", "τxy (Pa)"]
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = BORDER
        labels = ["N1 (--)", "N2 (+-)", "N3 (++)", "N4 (-+)"]
        r = 4
        for el_idx, el_res in enumerate(result.elements, start=1):
            for k, (eps, sig) in enumerate(
                zip(el_res.strains_at_corners, el_res.stresses_at_corners)
            ):
                ws.cell(row=r, column=1, value=el_idx).border = BORDER
                ws.cell(row=r, column=2, value=labels[k]).border = BORDER
                for c, v in enumerate(
                    [eps[0], eps[1], eps[2], sig[0], sig[1], sig[2]], start=3
                ):
                    cell = ws.cell(row=r, column=c, value=float(v))
                    cell.number_format = NUM_FMT
                    cell.border = BORDER
                r += 1
        for col_letter in "ABCDEFGH":
            ws.column_dimensions[col_letter].width = 16

    # ===== Hojas siguientes: matrices clave =====
    # K_global y K^e
    matrices_to_export = {}
    for st in procedure.steps:
        for mname, M in st.matrices.items():
            mname_lower = mname.lower()
            if ("k global" in mname_lower or "k^e" in mname_lower
                    or "k_ff" in mname_lower or "k_pen" in mname_lower
                    or mname_lower.startswith("3) d resultante")):
                # Limitar nombres a 31 caracteres
                safe = _safe_sheet_name(mname)
                if safe not in matrices_to_export:
                    matrices_to_export[safe] = M
    for sheet_name, M in matrices_to_export.items():
        try:
            ws = wb.create_sheet(sheet_name)
            ws.cell(row=1, column=1, value=sheet_name).font = TITLE_FONT
            _write_matrix(ws, sheet_name, M, 3)
        except Exception:
            pass

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))


def export_to_excel(procedure: Procedure, path: Path) -> None:
    wb = Workbook()
    # Eliminar hoja por defecto
    default = wb.active
    wb.remove(default)

    # Hoja resumen
    ws_summary = wb.create_sheet("Resumen")
    ws_summary.cell(row=1, column=1, value="Análisis MEF — Resumen").font = TITLE_FONT
    ws_summary.cell(row=3, column=1, value="Paso").font = HEADER_FONT
    ws_summary.cell(row=3, column=2, value="Descripción").font = HEADER_FONT
    for i, st in enumerate(procedure.steps, start=4):
        ws_summary.cell(row=i, column=1, value=st.title)
        ws_summary.cell(row=i, column=2, value=st.description)
        ws_summary.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws_summary.column_dimensions["A"].width = 40
    ws_summary.column_dimensions["B"].width = 80

    # Una hoja por paso
    for i, step in enumerate(procedure.steps, start=1):
        sheet_name = _safe_sheet_name(f"{i:02d} {step.title}")
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=1, column=1, value=step.title).font = TITLE_FONT
        ws.cell(row=2, column=1, value=step.description).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[2].height = 40
        ws.column_dimensions["A"].width = 22
        for col_letter in "BCDEFGH":
            ws.column_dimensions[col_letter].width = 18

        row = 4
        # Tablas
        for name, df in step.tables.items():
            ws.cell(row=row, column=1, value=name).font = Font(bold=True)
            row += 1
            row = _write_dataframe(ws, df, row)
        # Matrices
        for name, M in step.matrices.items():
            row = _write_matrix(ws, name, M, row)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
